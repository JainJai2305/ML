from sys import intern
from openpyxl.workbook import workbook
import requests
from openpyxl import load_workbook
from bs4 import BeautifulSoup

book= load_workbook('internships.xlsx')
ws=book.worksheets[0]
ws_tables =[]

for i in range(2,10) :ws['A'+str(i)]="Number of openings"
book.save('internships.xlsx')
